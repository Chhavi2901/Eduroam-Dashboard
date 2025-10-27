import re
import pandas as pd
import tempfile
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.chart import PieChart, BarChart, Reference
import numpy as np
from concurrent.futures import ThreadPoolExecutor
import threading


def extract_summary_tables_from_stream(stream, chunk_size=1_000_000):
    """
    Optimized stream-parser for very large .log files.
    - Uses faster regex patterns and optimized batch processing
    - Writes Access and F-TICKS entries to temp CSVs in larger chunks
    - Returns sampled DataFrames for visualization + paths to full CSVs.
    """

    # Optimized regex patterns with non-capturing groups where possible
    access_pattern = re.compile(
        r'(\w{3}\s\w{3}\s\d+\s\d{2}:\d{2}:\d{2}\s\d{4}):\s*'
        r'(Access-(?:Accept|Reject))\s+for\s+user\s+([\w@.\-]+)'
        r'\s+from\s+([\w._\-]+)\s+to\s+([\w._\-]+)\s*\(([\d.]+)\)',
        re.IGNORECASE
    )
    fticks_pattern = re.compile(
        r'(\w{3}\s\w{3}\s\d+\s\d{2}:\d{2}:\d{2}\s\d{4}):\s*'
        r'F-TICKS/eduroam/1\.0#REALM=([^#]*)#VISCOUNTRY=([^#]*)#'
        r'VISINST=([^#]*)#CSI=([^#]*)#RESULT=([^#]*)#',
        re.IGNORECASE
    )

    tmp_access = tempfile.NamedTemporaryFile(delete=False, suffix="_access.csv", mode="w", encoding="utf-8", buffering=8192)
    tmp_fticks = tempfile.NamedTemporaryFile(delete=False, suffix="_fticks.csv", mode="w", encoding="utf-8", buffering=8192)
    tmp_access.write("Timestamp,Event,Username,Source,Destination,ServerIP\n")
    tmp_fticks.write("Timestamp,REALM,VISCOUNTRY,VISINST,CSI,RESULT,Reason\n")

    access_batch, fticks_batch = [], []
    sample_access, sample_fticks = [], []
    write_lock = threading.Lock()

    def flush_batches():
        with write_lock:
            if access_batch:
                # Use faster CSV writing
                access_df = pd.DataFrame(access_batch)
                access_df.to_csv(tmp_access, header=False, index=False, mode='a')
                access_batch.clear()
            if fticks_batch:
                fticks_df = pd.DataFrame(fticks_batch)
                fticks_df.to_csv(tmp_fticks, header=False, index=False, mode='a')
                fticks_batch.clear()

    # Pre-compile string checks for faster processing
    access_marker = b"Access-"
    fticks_marker = b"F-TICKS"
    sample_limit = 50_000  # Reduced sample size for faster processing
    
    for i, raw_line in enumerate(stream):
        # Skip empty lines early
        if not raw_line or len(raw_line) < 20:
            continue
            
        # Fast byte-level checks before string conversion
        if access_marker in raw_line:
            try:
                line = raw_line.decode("utf-8", errors="ignore").strip()
                m = access_pattern.search(line)
                if m:
                    rec = {
                        'Timestamp': m.group(1),
                        'Event': m.group(2),
                        'Username': m.group(3),
                        'Source': m.group(4),
                        'Destination': m.group(5),
                        'ServerIP': m.group(6),
                    }
                    access_batch.append(rec)
                    if len(sample_access) < sample_limit:
                        sample_access.append(rec)
            except Exception:
                continue

        elif fticks_marker in raw_line:
            try:
                line = raw_line.decode("utf-8", errors="ignore").strip()
                m = fticks_pattern.search(line)
                if m:
                    res = m.group(6).strip()
                    rec = {
                        'Timestamp': m.group(1),
                        'REALM': m.group(2).strip(),
                        'VISCOUNTRY': m.group(3).strip(),
                        'VISINST': m.group(4).strip(),
                        'CSI': m.group(5).strip(),
                        'RESULT': res,
                        'Reason': "Authentication successful" if res.upper() == "OK" else "Authentication failed",
                    }
                    fticks_batch.append(rec)
                    if len(sample_fticks) < sample_limit:
                        sample_fticks.append(rec)
            except Exception:
                continue

        # Flush batches more frequently for better memory management
        if (i + 1) % chunk_size == 0:
            flush_batches()

    flush_batches()
    tmp_access.close()
    tmp_fticks.close()

    # Faster DataFrame creation with optimized data types
    if sample_access:
        df_access = pd.DataFrame(sample_access)
        # Use faster datetime parsing
        df_access["Timestamp"] = pd.to_datetime(df_access["Timestamp"], errors="coerce", format="%a %b %d %H:%M:%S %Y")
    else:
        df_access = pd.DataFrame()
        
    if sample_fticks:
        df_fticks = pd.DataFrame(sample_fticks)
        # Use faster datetime parsing
        df_fticks["Timestamp"] = pd.to_datetime(df_fticks["Timestamp"], errors="coerce", format="%a %b %d %H:%M:%S %Y")
    else:
        df_fticks = pd.DataFrame()

    return df_access, df_fticks, tmp_access.name, tmp_fticks.name


def parse_and_generate_excel(log_data: str) -> BytesIO:
    """
    Parses raw log text and generates an Excel file with charts.
    (Used for manual Excel export.)
    """
    df_access, df_fticks, _, _ = extract_summary_tables_from_stream(BytesIO(log_data.encode("utf-8")))

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_access.to_excel(writer, sheet_name="AccessLogs", index=False)
        df_fticks.to_excel(writer, sheet_name="FTicksLogs", index=False)

    output.seek(0)
    wb = load_workbook(output)

    ws_fticks = wb["FTicksLogs"]
    if not df_fticks.empty:
        pie = PieChart()
        pie.title = "Authentication Result Distribution"
        data_ref = Reference(ws_fticks, min_col=6, min_row=1, max_row=len(df_fticks) + 1)
        pie.add_data(data_ref, titles_from_data=True)
        ws_fticks.add_chart(pie, "H2")

        df_visinst = df_fticks["VISINST"].value_counts().reset_index()
        df_visinst.columns = ["VISINST", "Count"]
        chart_ws = wb.create_sheet("ChartData")
        chart_ws.append(["VISINST", "Count"])
        for row in df_visinst.values:
            chart_ws.append(list(row))

        bar = BarChart()
        bar.title = "Top VISINSTs"
        data_ref = Reference(chart_ws, min_col=2, min_row=1, max_row=len(df_visinst) + 1)
        label_ref = Reference(chart_ws, min_col=1, min_row=2, max_row=len(df_visinst) + 1)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(label_ref)
        ws_fticks.add_chart(bar, "H20")

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output
